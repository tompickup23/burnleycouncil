#!/usr/bin/env python3
"""
hmo_etl.py — HMO (Houses in Multiple Occupation) register ETL for Lancashire councils.

Sources (tiered by availability):
  TIER 1 — Direct downloads: Chorley (XLSX), Lancaster (PDF), Ribble Valley (PDF)
  TIER 2 — Web portals: Preston (ASP.NET, 307 records), Blackpool, Burnley
  TIER 3 — Planning applications: Extract HMO-related apps from planning.json
  TIER 4 — FOI required: Hyndburn, Pendle, Fylde, Wyre, Blackburn, West Lancashire

Outputs per council: hmo.json with:
  - register: licensed HMO list (from council register where available)
  - planning_hmos: HMO-related planning applications
  - summary: ward-level counts, totals, density metrics
  - meta: source, fetch date, coverage

Usage:
  python3 hmo_etl.py --council preston
  python3 hmo_etl.py --all
"""

import argparse
import json
import os
import re
import sys
import tempfile
import time
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, '..', 'data')

# ── Council metadata ─────────────────────────────────────────────────────────
COUNCIL_HMO_SOURCES = {
    'preston': {
        'type': 'web_portal_preston',
        'url': 'https://selfservice.preston.gov.uk/service/publicregister/hmosearch.aspx',
        'register_name': 'Preston City Council HMO Public Register',
    },
    'blackpool': {
        'type': 'web_portal',
        'url': 'https://selfservice.blackpool.gov.uk/Landlords/housinglicensing/PublicRegister',
        'register_name': 'Blackpool Council Housing Licensing Public Register',
    },
    'chorley': {
        'type': 'download_xlsx',
        'url': 'https://chorley.gov.uk/downloads/file/1473/hmo-register',
        'register_name': 'Chorley Council HMO Register',
    },
    'lancaster': {
        'type': 'download_pdf',
        'url': 'https://www.lancaster.gov.uk/assets/attach/14038/HMO-Licensing-Register.-Nov.-24.pdf',
        'register_name': 'Lancaster City Council HMO Register (Nov 2024)',
    },
    'ribble_valley': {
        'type': 'download_pdf',
        'url': 'https://www.ribblevalley.gov.uk/downloads/file/2415/register-of-licensed-houses-in-multiple-occupation',
        'register_name': 'Ribble Valley HMO Register',
    },
    'south_ribble': {
        'type': 'web_portal',
        'url': 'https://southribble.gov.uk/environmental-health/hmo/4',
        'register_name': 'South Ribble HMO Register',
    },
    'burnley': {
        'type': 'web_portal',
        'url': 'https://propertylicensing.burnley.gov.uk/online-application/public-register/',
        'register_name': 'Burnley Property Licensing Public Register',
    },
    # TIER 3 — planning-only (no public register)
    'hyndburn':        {'type': 'planning_only', 'register_name': None},
    'pendle':          {'type': 'planning_only', 'register_name': None},
    'rossendale':      {'type': 'planning_only', 'register_name': None},
    'fylde':           {'type': 'planning_only', 'register_name': None},
    'wyre':            {'type': 'planning_only', 'register_name': None},
    'blackburn':       {'type': 'planning_only', 'register_name': None},
    'west_lancashire': {'type': 'planning_only', 'register_name': None},
}

# ── Utility functions ────────────────────────────────────────────────────────

def fetch_url(url, timeout=60, retries=3):
    """Fetch URL with retries."""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
                'Accept': '*/*',
            })
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as e:
            if attempt < retries - 1:
                wait = 10 * (attempt + 1)
                print(f"    ⚠ Attempt {attempt+1} failed: {e}. Retrying in {wait}s...")
                time.sleep(wait)
            else:
                print(f"    ✗ Failed after {retries} attempts: {e}")
                return None
    return None


def fetch_url_with_post(url, data, timeout=60, cookies=None):
    """POST to URL and return response bytes."""
    try:
        encoded = urllib.parse.urlencode(data).encode('utf-8')
        req = urllib.request.Request(url, data=encoded, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Content-Type': 'application/x-www-form-urlencoded',
        })
        if cookies:
            req.add_header('Cookie', cookies)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read(), resp.headers.get('Set-Cookie', '')
    except Exception as e:
        print(f"    ⚠ POST failed: {e}")
        return None, ''


def bulk_geocode_postcodes(postcodes):
    """Bulk geocode postcodes via postcodes.io (100 per request)."""
    results = {}
    unique_pcs = list(set(pc.strip().upper() for pc in postcodes if pc and pc.strip()))

    for i in range(0, len(unique_pcs), 100):
        batch = unique_pcs[i:i+100]
        payload = json.dumps({'postcodes': batch}).encode('utf-8')
        try:
            req = urllib.request.Request(
                'https://api.postcodes.io/postcodes',
                data=payload,
                headers={
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode('utf-8'))
                for item in data.get('result', []):
                    if item.get('result'):
                        r = item['result']
                        results[item['query'].replace(' ', '').upper()] = {
                            'lat': r.get('latitude'),
                            'lng': r.get('longitude'),
                            'ward': r.get('admin_ward'),
                            'district': r.get('admin_district'),
                            'lsoa': r.get('lsoa'),
                            'msoa': r.get('msoa'),
                        }
        except Exception as e:
            print(f"    ⚠ Bulk geocode batch failed: {e}")
        if i + 100 < len(unique_pcs):
            time.sleep(1)
    return results


def extract_postcode(address):
    """Extract UK postcode from an address string."""
    if not address:
        return None
    m = re.search(r'([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})', address.upper())
    return m.group(1) if m else None


# ── Preston scraper (ASP.NET with __doPostBack pagination) ───────────────────

def scrape_preston_hmos():
    """Scrape all pages of Preston's ASP.NET HMO register.

    Page structure (confirmed from HTML):
      <div id="applicationcontainer"><div id="applicationdetails">
        <b>Licence number: </b>HMO0001<br/>
        <b>Address: </b>9 Osborne Street, Preston, PR1 8PN<br/>
        <b>Licence holder: </b>Mr Andrew Halsall<br/>
        <div id="divnumber1" style="display:none;">
          <b>Commencement date: </b>05/07/2021<br/>
          <b>Expiry date: </b>05/07/2026<br/>
          <b>Maximum occupancy: </b>6<br/>
          <b>Number of floors: </b>3<br/>
          <b>Number of households: </b>6<br/>
        </div>
      </div></div>
      <hr/>

    Pagination: __doPostBack('ctl00$MainContent$dlPager$ctl{NN}$lnkPageNo','')
    Pages: 10 per page, "307 records found", 31 pages.
    """
    print("    Scraping Preston HMO register...")
    base_url = 'https://selfservice.preston.gov.uk/service/publicregister/hmosearch.aspx'

    html = fetch_url(base_url)
    if not html:
        return []

    html_text = html.decode('utf-8', errors='replace')
    all_hmos = parse_preston_page(html_text)
    print(f"    Page 1: {len(all_hmos)} entries")

    # Detect total pages
    m = re.search(r'(\d+)\s+records?\s+found', html_text)
    total_records = int(m.group(1)) if m else 0
    total_pages = (total_records + 9) // 10 if total_records else 1
    print(f"    {total_records} records, {total_pages} pages")

    for page_num in range(2, total_pages + 1):
        # Extract ASP.NET form state
        viewstate = extract_asp_field(html_text, '__VIEWSTATE')
        viewstate_gen = extract_asp_field(html_text, '__VIEWSTATEGENERATOR')
        event_val = extract_asp_field(html_text, '__EVENTVALIDATION')

        if not viewstate:
            print(f"    ⚠ No viewstate on page {page_num-1}, stopping")
            break

        # Preston uses 0-indexed pager controls: ctl00, ctl01, ctl02...
        # Page 2 = ctl01, Page 3 = ctl02, etc.
        pager_idx = page_num - 1
        event_target = f'ctl00$MainContent$dlPager$ctl{pager_idx:02d}$lnkPageNo'

        form_data = {
            '__VIEWSTATE': viewstate,
            '__VIEWSTATEGENERATOR': viewstate_gen,
            '__EVENTVALIDATION': event_val,
            '__EVENTTARGET': event_target,
            '__EVENTARGUMENT': '',
        }

        resp_bytes, _ = fetch_url_with_post(base_url, form_data)
        if not resp_bytes:
            # Try alternate pager naming
            event_target = f'ctl00$MainContent$dlPager2$ctl{pager_idx:02d}$lnkPageNo'
            form_data['__EVENTTARGET'] = event_target
            resp_bytes, _ = fetch_url_with_post(base_url, form_data)
            if not resp_bytes:
                print(f"    ⚠ Page {page_num} failed, stopping")
                break

        html_text = resp_bytes.decode('utf-8', errors='replace')
        page_hmos = parse_preston_page(html_text)
        all_hmos.extend(page_hmos)

        if page_num % 5 == 0 or page_num == total_pages:
            print(f"    Page {page_num}/{total_pages}: +{len(page_hmos)} (total: {len(all_hmos)})")

        time.sleep(1.5)  # Be polite

    print(f"    ✓ Total: {len(all_hmos)} HMOs from Preston register")
    return all_hmos


def extract_asp_field(html, field_name):
    """Extract ASP.NET hidden field value from HTML."""
    pattern = rf'id="{re.escape(field_name)}"[^>]*value="([^"]*)"'
    m = re.search(pattern, html, re.IGNORECASE)
    if m:
        return m.group(1)
    pattern = rf'name="{re.escape(field_name)}"[^>]*value="([^"]*)"'
    m = re.search(pattern, html, re.IGNORECASE)
    return m.group(1) if m else ''


def parse_preston_page(html):
    """Parse HMO entries from Preston's HTML using the exact known structure.

    Each entry follows: <b>Label: </b>Value<br/> pattern inside applicationdetails divs.
    """
    hmos = []

    # Split on <hr/> which separates each entry
    # The data sits inside MainContent_lblDetails span
    label_match = re.search(r'id="MainContent_lblDetails"[^>]*>(.*?)</span>', html, re.DOTALL)
    if not label_match:
        return hmos

    content = label_match.group(1)

    # Split entries by <hr/>
    entries = re.split(r'<hr\s*/?\s*>', content)

    for entry in entries:
        hmo = {}

        # Extract all <b>Label: </b>Value pairs
        pairs = re.findall(r'<b>\s*(.*?)\s*</b>\s*(.*?)(?=<b>|<div|<a|</div|$)', entry, re.DOTALL)

        for label_raw, value_raw in pairs:
            label = label_raw.strip().rstrip(':').lower()
            # Strip HTML tags from value
            value = re.sub(r'<[^>]+>', '', value_raw).strip().rstrip('<').strip()
            if not label or not value:
                continue

            if 'licence number' in label:
                hmo['licence_number'] = value
            elif 'address' in label:
                hmo['address'] = value
            elif 'licence holder' in label or 'holder' in label:
                hmo['licence_holder'] = value
            elif 'commencement' in label:
                hmo['commencement_date'] = value
            elif 'expiry' in label or 'expir' in label:
                hmo['expiry_date'] = value
            elif 'maximum occupancy' in label or 'occupan' in label:
                try:
                    hmo['max_occupants'] = int(re.search(r'\d+', value).group())
                except (AttributeError, ValueError):
                    pass
            elif 'floors' in label or 'floor' in label:
                try:
                    hmo['floors'] = int(re.search(r'\d+', value).group())
                except (AttributeError, ValueError):
                    pass
            elif 'households' in label or 'household' in label:
                try:
                    hmo['households'] = int(re.search(r'\d+', value).group())
                except (AttributeError, ValueError):
                    pass

        if hmo.get('address'):
            hmo['postcode'] = extract_postcode(hmo['address'])
            hmos.append(hmo)

    return hmos


# ── Chorley XLSX parser ──────────────────────────────────────────────────────

def parse_chorley_xlsx():
    """Download and parse Chorley's HMO register XLSX."""
    print("    Downloading Chorley HMO register (XLSX)...")
    url = 'https://chorley.gov.uk/downloads/file/1473/hmo-register'

    data = fetch_url(url)
    if not data:
        return []

    # Write to temp file for openpyxl
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.write(data)
    tmp.close()

    hmos = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]

        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or '').strip().lower() for c in row]
                continue

            if not any(row):
                continue

            cells = dict(zip(headers, row))
            address = str(cells.get('address of licensed hmo', '') or '').replace('\n', ', ').strip()
            if not address:
                continue

            hmo = {
                'address': address,
                'postcode': extract_postcode(address),
            }

            # Max occupants
            occ = cells.get('maximum number of occupants', cells.get('maximum number of occupants ', ''))
            if occ:
                try:
                    hmo['max_occupants'] = int(occ)
                except (ValueError, TypeError):
                    pass

            # Dates
            issue = cells.get('licence issue date', '')
            expiry = cells.get('licence expiry date', '')
            if issue:
                hmo['commencement_date'] = str(issue)[:10] if hasattr(issue, 'strftime') else str(issue)
            if expiry:
                hmo['expiry_date'] = str(expiry)[:10] if hasattr(expiry, 'strftime') else str(expiry)

            # Rooms
            rooms = cells.get('number of rooms – sleeping accomodation', cells.get('number of rooms – sleeping accommodation', ''))
            if rooms:
                try:
                    hmo['bedrooms'] = int(rooms)
                except (ValueError, TypeError):
                    pass

            storeys = cells.get('number of storeys', '')
            if storeys:
                try:
                    hmo['floors'] = int(storeys)
                except (ValueError, TypeError):
                    pass

            desc = cells.get('description of hmo', '')
            if desc:
                hmo['description'] = str(desc).strip()

            uprn = cells.get('uprn', '')
            if uprn:
                try:
                    hmo['uprn'] = int(uprn)
                except (ValueError, TypeError):
                    hmo['uprn'] = str(uprn)

            hmos.append(hmo)

        wb.close()
    except ImportError:
        print("    ⚠ openpyxl not installed — cannot parse XLSX")
    except Exception as e:
        print(f"    ⚠ Error parsing Chorley XLSX: {e}")
    finally:
        os.unlink(tmp.name)

    print(f"    ✓ {len(hmos)} HMOs from Chorley register (XLSX)")
    return hmos


# ── Lancaster / Ribble Valley PDF parser ─────────────────────────────────────

def parse_hmo_pdf(url, council_name):
    """Download and parse HMO register PDF using pdfplumber."""
    print(f"    Downloading {council_name} HMO register (PDF)...")

    data = fetch_url(url, timeout=120)
    if not data:
        return []

    tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    tmp.write(data)
    tmp.close()

    hmos = []
    try:
        import pdfplumber

        with pdfplumber.open(tmp.name) as pdf:
            all_text = ''
            for page in pdf.pages:
                # Try extracting tables first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if not row:
                                continue
                            # Look for rows with postcodes (HMO entries)
                            row_text = ' '.join(str(c or '') for c in row)
                            pc = extract_postcode(row_text)
                            if pc:
                                hmo = {
                                    'address': row_text.strip(),
                                    'postcode': pc,
                                }
                                # Try to extract max occupants
                                nums = re.findall(r'\b(\d{1,3})\b', row_text)
                                for n in nums:
                                    n_int = int(n)
                                    if 2 <= n_int <= 100:
                                        hmo['max_occupants'] = n_int
                                        break
                                hmos.append(hmo)

                # Also extract text for non-tabular PDFs
                text = page.extract_text() or ''
                all_text += text + '\n'

            # If no table-based results, try text-based extraction
            if not hmos and all_text:
                # Split by address-like patterns
                # Look for lines containing postcodes
                for line in all_text.split('\n'):
                    line = line.strip()
                    pc = extract_postcode(line)
                    if pc and len(line) > 15:
                        hmo = {
                            'address': line,
                            'postcode': pc,
                        }
                        hmos.append(hmo)

    except ImportError:
        print("    ⚠ pdfplumber not installed — cannot parse PDF")
    except Exception as e:
        print(f"    ⚠ Error parsing {council_name} PDF: {e}")
    finally:
        os.unlink(tmp.name)

    # Deduplicate by postcode
    seen = set()
    unique_hmos = []
    for hmo in hmos:
        key = (hmo.get('address', ''), hmo.get('postcode', ''))
        if key not in seen:
            seen.add(key)
            unique_hmos.append(hmo)

    print(f"    ✓ {len(unique_hmos)} HMOs from {council_name} register (PDF)")
    return unique_hmos


# ── Planning-based HMO extraction ────────────────────────────────────────────

HMO_PATTERNS = [
    re.compile(r'\bHMO\b', re.IGNORECASE),
    re.compile(r'house[s]?\s+in\s+multiple\s+occupation', re.IGNORECASE),
    re.compile(r'multiple\s+occupation', re.IGNORECASE),
    re.compile(r'\bC3\b.*\bC4\b', re.IGNORECASE),
    re.compile(r'\bC4\b.*\bHMO\b', re.IGNORECASE),
    re.compile(r'use\s+class\s+C4', re.IGNORECASE),
    re.compile(r'bedsit', re.IGNORECASE),
]


def extract_hmos_from_planning(council_id):
    """Extract HMO-related applications from planning.json."""
    planning_path = os.path.join(DATA_DIR, council_id, 'planning.json')
    if not os.path.exists(planning_path):
        return []

    with open(planning_path) as f:
        planning = json.load(f)

    apps = planning.get('applications', [])
    hmo_apps = []
    seen_uids = set()

    for app in apps:
        uid = app.get('uid', '')
        if uid in seen_uids:
            continue

        desc = (app.get('description') or '').strip()
        if not desc:
            continue

        is_hmo = any(p.search(desc) for p in HMO_PATTERNS)
        if not is_hmo:
            continue

        seen_uids.add(uid)
        hmo_apps.append({
            'uid': uid,
            'address': app.get('address', ''),
            'postcode': app.get('postcode') or extract_postcode(app.get('address', '')),
            'ward': app.get('ward', ''),
            'lat': app.get('lat'),
            'lng': app.get('lng'),
            'description': desc,
            'type': app.get('type', ''),
            'state': app.get('state', ''),
            'start_date': app.get('start_date', ''),
            'decided_date': app.get('decided_date'),
            'source': 'planning_application',
        })

    return hmo_apps


# ── Ward aggregation ─────────────────────────────────────────────────────────

def aggregate_by_ward(register_hmos, planning_hmos, geocode_results):
    """Aggregate HMO data by ward."""
    ward_counts = defaultdict(lambda: {'register': 0, 'planning_apps': 0, 'postcodes': set()})

    for hmo in register_hmos:
        ward = hmo.get('ward', '')
        if not ward:
            pc = (hmo.get('postcode') or '').replace(' ', '').upper()
            geo = geocode_results.get(pc, {})
            ward = geo.get('ward', 'Unknown')
        ward_counts[ward]['register'] += 1
        pc = hmo.get('postcode', '')
        if pc:
            ward_counts[ward]['postcodes'].add(pc.replace(' ', '').upper())

    for hmo in planning_hmos:
        ward = hmo.get('ward', '')
        if not ward:
            pc = (hmo.get('postcode') or '').replace(' ', '').upper()
            geo = geocode_results.get(pc, {})
            ward = geo.get('ward', 'Unknown')
        if ward:
            ward_counts[ward]['planning_apps'] += 1

    result = {}
    for ward, data in sorted(ward_counts.items(), key=lambda x: -(x[1]['register'] + x[1]['planning_apps'])):
        if ward == 'Unknown' and data['register'] == 0 and data['planning_apps'] == 0:
            continue
        result[ward] = {
            'licensed_hmos': data['register'],
            'planning_applications': data['planning_apps'],
            'total': data['register'] + data['planning_apps'],
            'postcodes': sorted(data['postcodes']),
        }

    return result


def load_ward_populations(council_id):
    """Load ward populations from demographics.json."""
    demo_path = os.path.join(DATA_DIR, council_id, 'demographics.json')
    if not os.path.exists(demo_path):
        return {}
    try:
        with open(demo_path) as f:
            demo = json.load(f)
        pops = {}
        wards = demo if isinstance(demo, dict) else {}
        for ward_code, wd in wards.items():
            ward_name = wd.get('ward_name', '')
            total = wd.get('total_population') or wd.get('population', {}).get('total')
            if ward_name and total:
                pops[ward_name] = total
        return pops
    except Exception:
        return {}


# ── Main processing ──────────────────────────────────────────────────────────

def process_council(council_id):
    """Process HMO data for a single council."""
    source = COUNCIL_HMO_SOURCES.get(council_id)
    if not source:
        print(f"  ✗ No HMO source config for {council_id}")
        return False

    council_dir = os.path.join(DATA_DIR, council_id)
    if not os.path.exists(council_dir):
        print(f"  ✗ No data directory for {council_id}")
        return False

    print(f"\n{'='*60}")
    print(f"  Processing HMOs: {council_id}")
    print(f"{'='*60}")

    register_hmos = []
    source_type = source['type']
    source_note = ''

    # ── TIER 1: Direct downloads ─────────────────────────────────────────
    if source_type == 'download_xlsx' and council_id == 'chorley':
        register_hmos = parse_chorley_xlsx()
        source_note = f"Downloaded from {source['register_name']} (XLSX)"

    elif source_type == 'download_pdf':
        register_hmos = parse_hmo_pdf(source['url'], source.get('register_name', council_id))
        source_note = f"Downloaded from {source['register_name']} (PDF)"

    # ── TIER 2: Web portals ──────────────────────────────────────────────
    elif source_type == 'web_portal_preston':
        register_hmos = scrape_preston_hmos()
        source_note = f"Scraped from {source['register_name']}"

    elif source_type == 'web_portal':
        print(f"    ℹ {council_id}: Register at {source['url']}")
        print(f"    ℹ Web portal — scraper not yet implemented")
        source_note = f"Register available at {source['url']} — scraper pending"

    # ── TIER 3: Planning only ────────────────────────────────────────────
    elif source_type == 'planning_only':
        source_note = 'No public register found — FOI required. Using planning applications only.'

    # ── Extract from planning data (all councils) ────────────────────────
    print(f"  Extracting HMOs from planning applications...")
    planning_hmos = extract_hmos_from_planning(council_id)
    print(f"    ✓ {len(planning_hmos)} HMO-related planning applications found")

    if not register_hmos and not planning_hmos:
        print(f"  ⚠ No HMO data found for {council_id}")
        output = {
            'meta': {
                'council_id': council_id,
                'fetched': datetime.now().strftime('%Y-%m-%d'),
                'source': source_note or 'No data available',
                'register_url': source.get('url'),
                'coverage': 'none',
            },
            'register': [],
            'planning_hmos': [],
            'summary': {
                'total_licensed': 0,
                'total_planning_apps': 0,
                'by_ward': {},
            },
        }
        out_path = os.path.join(council_dir, 'hmo.json')
        with open(out_path, 'w') as f:
            json.dump(output, f, separators=(',', ':'))
        print(f"  ✓ Written empty {out_path}")
        return True

    # ── Geocode postcodes to wards ───────────────────────────────────────
    all_postcodes = set()
    for hmo in register_hmos:
        pc = hmo.get('postcode')
        if pc:
            all_postcodes.add(pc)
    for hmo in planning_hmos:
        pc = hmo.get('postcode')
        if pc:
            all_postcodes.add(pc)

    geocode_results = {}
    if all_postcodes:
        print(f"  Geocoding {len(all_postcodes)} postcodes...")
        geocode_results = bulk_geocode_postcodes(list(all_postcodes))
        print(f"    ✓ {len(geocode_results)}/{len(all_postcodes)} geocoded")

        for hmo in register_hmos:
            pc = (hmo.get('postcode') or '').replace(' ', '').upper()
            geo = geocode_results.get(pc, {})
            if geo:
                hmo['ward'] = geo.get('ward', '')
                hmo['lat'] = geo.get('lat')
                hmo['lng'] = geo.get('lng')
                hmo['lsoa'] = geo.get('lsoa', '')

        for hmo in planning_hmos:
            if not hmo.get('ward'):
                pc = (hmo.get('postcode') or '').replace(' ', '').upper()
                geo = geocode_results.get(pc, {})
                if geo:
                    hmo['ward'] = geo.get('ward', '')

    # ── Aggregate by ward ────────────────────────────────────────────────
    by_ward = aggregate_by_ward(register_hmos, planning_hmos, geocode_results)

    # ── Ward population density ──────────────────────────────────────────
    ward_pops = load_ward_populations(council_id)
    for ward_name, ward_data in by_ward.items():
        pop = ward_pops.get(ward_name, 0)
        if pop > 0:
            ward_data['density_per_1000'] = round(ward_data['total'] / pop * 1000, 2)
            ward_data['population'] = pop

    # ── Build output ─────────────────────────────────────────────────────
    total_licensed = len(register_hmos)
    total_planning = len(planning_hmos)

    decided = [h for h in planning_hmos if h.get('state') in
               ('Permitted', 'Approved', 'Granted', 'Conditions', 'Refused', 'Rejected')]
    approved = [h for h in planning_hmos if h.get('state') in
                ('Permitted', 'Approved', 'Granted', 'Conditions')]
    refused = [h for h in planning_hmos if h.get('state') in ('Refused', 'Rejected')]

    # Postcode hotspots
    pc_counts = defaultdict(int)
    for hmo in register_hmos + planning_hmos:
        pc = hmo.get('postcode', '')
        if pc:
            outward = pc.split()[0] if ' ' in pc else pc[:4].strip()
            pc_counts[outward] += 1
    hotspots = sorted(pc_counts.items(), key=lambda x: -x[1])[:10]

    # Max occupants stats
    occupants = [h.get('max_occupants', 0) for h in register_hmos if h.get('max_occupants')]
    total_bed_spaces = sum(occupants) if occupants else None
    avg_occupants = round(sum(occupants) / len(occupants), 1) if occupants else None

    output = {
        'meta': {
            'council_id': council_id,
            'fetched': datetime.now().strftime('%Y-%m-%d'),
            'source': source_note,
            'register_url': source.get('url'),
            'register_name': source.get('register_name'),
            'coverage': 'register+planning' if register_hmos else 'planning_only',
        },
        'register': register_hmos,
        'planning_hmos': planning_hmos,
        'summary': {
            'total_licensed': total_licensed,
            'total_planning_apps': total_planning,
            'total_combined': total_licensed + total_planning,
            'total_bed_spaces': total_bed_spaces,
            'avg_occupants': avg_occupants,
            'planning_approved': len(approved),
            'planning_refused': len(refused),
            'planning_pending': total_planning - len(decided),
            'planning_approval_rate': round(len(approved) / len(decided), 3) if decided else None,
            'by_ward': {k: {kk: vv for kk, vv in v.items() if kk != 'postcodes'} for k, v in by_ward.items()},
            'hotspot_postcodes': [{'postcode': pc, 'count': c} for pc, c in hotspots],
            'top_ward': max(by_ward.items(), key=lambda x: x[1]['total'])[0] if by_ward else None,
        },
    }

    # ── Write output ─────────────────────────────────────────────────────
    out_path = os.path.join(council_dir, 'hmo.json')
    with open(out_path, 'w') as f:
        json.dump(output, f, separators=(',', ':'))

    size_kb = os.path.getsize(out_path) // 1024
    print(f"  ✓ Written {out_path} ({size_kb}KB)")
    print(f"    Licensed HMOs: {total_licensed} | Planning apps: {total_planning}")
    if total_bed_spaces:
        print(f"    Total bed spaces: {total_bed_spaces} | Avg occupants: {avg_occupants}")
    if by_ward:
        top_3 = list(by_ward.items())[:3]
        for ward, data in top_3:
            print(f"    Top ward: {ward} — {data['total']} HMOs")

    return True


def main():
    parser = argparse.ArgumentParser(description='HMO Register ETL for Lancashire')
    parser.add_argument('--council', default=None, help='Single council ID')
    parser.add_argument('--all', action='store_true', help='Process all councils')
    args = parser.parse_args()

    councils = []
    if args.all:
        councils = list(COUNCIL_HMO_SOURCES.keys())
    elif args.council:
        councils = [args.council]
    else:
        parser.print_help()
        sys.exit(1)

    print(f"HMO ETL — {len(councils)} council(s)")
    print(f"Sources: Council registers + PlanIt planning applications")

    success = 0
    for cid in councils:
        try:
            if process_council(cid):
                success += 1
        except Exception as e:
            print(f"  ✗ Error processing {cid}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*60}")
    print(f"  Done: {success}/{len(councils)} councils processed")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
