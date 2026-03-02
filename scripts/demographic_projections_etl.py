#!/usr/bin/env python3
"""
demographic_projections_etl.py — ONS 2022-based SNPP + Home Office asylum data.

Fetches:
1. ONS Sub-National Population Projections (2022-based) via Nomis API
   - Council-level projections 2022→2047 with age band breakdown
2. Home Office asylum/resettlement statistics by local authority
   - Asy_D11: Asylum seekers receiving support
   - Res_D01: Refugee resettlement

Output per council: burnley-council/data/{council_id}/demographic_projections.json
Output shared:      burnley-council/data/shared/demographic_projections_summary.json

Usage:
    python3 scripts/demographic_projections_etl.py                    # All councils
    python3 scripts/demographic_projections_etl.py --council burnley  # Single council
    python3 scripts/demographic_projections_etl.py --skip-asylum      # Skip asylum data (Nomis only)
"""

import csv
import io
import json
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "burnley-council" / "data"

NOMIS_BASE = "https://www.nomisweb.co.uk/api/v01"
SNPP_DATASET = "NM_2006_1"  # 2022-based SNPP

# Age category codes in Nomis SNPP dataset
AGE_ALL = 200       # All Ages
AGE_0_15 = 201      # Aged 0 to 15
AGE_16_64 = 203     # Aged 16 to 64
AGE_65_PLUS = 209   # Aged 65+

PROJECTION_YEARS = [2022, 2027, 2032, 2037, 2042, 2047]

# Home Office asylum data URLs (Dec 2025 release)
ASYLUM_URL = "https://assets.publishing.service.gov.uk/media/69959470047739fe61889d47/support-local-authority-datasets-dec-2025.xlsx"
RESETTLEMENT_URL = "https://assets.publishing.service.gov.uk/media/69959395bfdab2546272bf06/resettlement-local-authority-datasets-dec-2025.xlsx"

# ONS codes for all 15 Lancashire councils
COUNCILS = {
    "burnley": {"ons": "E07000117", "name": "Burnley"},
    "hyndburn": {"ons": "E07000120", "name": "Hyndburn"},
    "pendle": {"ons": "E07000122", "name": "Pendle"},
    "rossendale": {"ons": "E07000125", "name": "Rossendale"},
    "lancaster": {"ons": "E07000121", "name": "Lancaster"},
    "ribble_valley": {"ons": "E07000124", "name": "Ribble Valley"},
    "chorley": {"ons": "E07000118", "name": "Chorley"},
    "south_ribble": {"ons": "E07000126", "name": "South Ribble"},
    "blackpool": {"ons": "E06000009", "name": "Blackpool"},
    "west_lancashire": {"ons": "E07000127", "name": "West Lancashire"},
    "blackburn": {"ons": "E06000008", "name": "Blackburn with Darwen"},
    "wyre": {"ons": "E07000128", "name": "Wyre"},
    "preston": {"ons": "E07000123", "name": "Preston"},
    "fylde": {"ons": "E07000119", "name": "Fylde"},
    "lancashire_cc": {"ons": "E10000017", "name": "Lancashire"},
}

# ---- Nomis API helpers ----

def fetch_csv(url, retries=3, delay=2):
    """Fetch CSV from URL with retry logic."""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "AI-DOGE-ETL/1.0 (Lancashire council transparency platform)"
            })
            with urllib.request.urlopen(req, timeout=60) as resp:
                return resp.read().decode("utf-8")
        except (urllib.error.URLError, TimeoutError) as e:
            if attempt < retries - 1:
                print(f"  Retry {attempt + 1}/{retries} after error: {e}", file=sys.stderr)
                time.sleep(delay)
            else:
                raise
    return ""


def fetch_snpp(ons_code):
    """Fetch SNPP projections for a council from Nomis API."""
    age_codes = f"{AGE_ALL},{AGE_0_15},{AGE_16_64},{AGE_65_PLUS}"
    years = ",".join(str(y) for y in PROJECTION_YEARS)

    url = (
        f"{NOMIS_BASE}/dataset/{SNPP_DATASET}.data.csv"
        f"?geography={ons_code}"
        f"&gender=0"  # Total (both sexes)
        f"&c_age={age_codes}"
        f"&projected_year={years}"
        f"&measures=20100"
        f"&select=geography_code,geography_name,projected_year_name,c_age_name,obs_value"
    )

    csv_text = fetch_csv(url)
    reader = csv.DictReader(io.StringIO(csv_text))

    # Parse into structured data
    projections = {}  # year -> {total, 0-15, 16-64, 65+}
    for row in reader:
        year = row.get("PROJECTED_YEAR_NAME", "").strip()
        age = row.get("C_AGE_NAME", "").strip()
        value = int(float(row.get("OBS_VALUE", 0)))

        if year not in projections:
            projections[year] = {}

        if "All Ages" in age:
            projections[year]["total"] = value
        elif "0 to 15" in age:
            projections[year]["0-15"] = value
        elif "16 to 64" in age:
            projections[year]["16-64"] = value
        elif "65+" in age:
            projections[year]["65+"] = value

    return projections


def compute_metrics(projections):
    """Compute derived metrics from projections."""
    result = {
        "population_projections": {},
        "age_projections": {},
        "dependency_ratio_projection": {},
        "working_age_pct_projection": {},
    }

    for year, data in sorted(projections.items()):
        total = data.get("total", 0)
        young = data.get("0-15", 0)
        working = data.get("16-64", 0)
        elderly = data.get("65+", 0)

        result["population_projections"][year] = total
        result["age_projections"][year] = {
            "0-15": young,
            "16-64": working,
            "65+": elderly,
        }

        # Dependency ratio: (0-15 + 65+) / (16-64) * 100
        if working > 0:
            dep_ratio = round(((young + elderly) / working) * 100, 1)
            result["dependency_ratio_projection"][year] = dep_ratio

        # Working age percentage
        if total > 0:
            result["working_age_pct_projection"][year] = round((working / total) * 100, 1)

    # Growth rate: base year to end year
    years = sorted(result["population_projections"].keys())
    if len(years) >= 2:
        base_pop = result["population_projections"][years[0]]
        end_pop = result["population_projections"][years[-1]]
        if base_pop > 0:
            result["growth_rate_pct"] = round(((end_pop - base_pop) / base_pop) * 100, 1)
        else:
            result["growth_rate_pct"] = 0
    else:
        result["growth_rate_pct"] = 0

    return result


# ---- Home Office asylum data ----

def fetch_asylum_data(skip=False):
    """Download and parse Home Office asylum Excel data (Asy_D11).

    Format: flat data rows with columns:
      Date, Support Type, Region, Local Authority, LAD Code, Accommodation Type, People

    We aggregate by LAD code for the latest date, grouping by accommodation type.
    Returns dict keyed by ONS code: {seekers_supported, by_accommodation, trend}
    """
    if skip:
        return {}

    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed — skipping asylum data. Install with: pip3 install openpyxl", file=sys.stderr)
        return {}

    print("Downloading Home Office asylum data...", file=sys.stderr)
    asylum_data = {}

    try:
        req = urllib.request.Request(ASYLUM_URL, headers={"User-Agent": "AI-DOGE-ETL/1.0"})
        import tempfile
        with urllib.request.urlopen(req, timeout=120) as resp:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(resp.read())
                tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

        # Use Data_Asy_D11 sheet (flat data, not pivot)
        target_sheet = None
        for name in wb.sheetnames:
            if "data" in name.lower() and "asy_d11" in name.lower():
                target_sheet = name
                break

        if not target_sheet:
            wb.close()
            Path(tmp_path).unlink(missing_ok=True)
            print("  WARNING: Could not find Data_Asy_D11 sheet", file=sys.stderr)
            return {}

        ws = wb[target_sheet]

        # Columns: Date, Support Type, Region, Local Authority, LAD Code, Accommodation Type, People
        # Header is row 1 (0-indexed), data starts row 2
        # Find the latest date first, then aggregate
        all_rows = []
        dates_seen = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # Title row
                continue
            if i == 1:  # Header row — verify structure
                continue
            if not row or len(row) < 7:
                continue
            date_val = str(row[0]).strip() if row[0] else ""
            if not date_val or date_val.startswith("Asy"):
                continue
            lad_code = str(row[4]).strip() if row[4] else ""
            if not lad_code.startswith("E"):
                continue

            try:
                people = int(float(row[6])) if row[6] else 0
            except (ValueError, TypeError):
                people = 0

            all_rows.append({
                "date": date_val,
                "support_type": str(row[1]).strip() if row[1] else "",
                "la_name": str(row[3]).strip() if row[3] else "",
                "lad_code": lad_code,
                "accommodation": str(row[5]).strip() if row[5] else "",
                "people": people,
            })
            dates_seen.add(date_val)

        # Find latest date
        if not dates_seen:
            wb.close()
            Path(tmp_path).unlink(missing_ok=True)
            print("  WARNING: No valid dates found in asylum data", file=sys.stderr)
            return {}

        latest_date = sorted(dates_seen)[-1]
        print(f"  Latest asylum data date: {latest_date}", file=sys.stderr)

        # Filter to latest date, aggregate by LA
        la_totals = {}  # lad_code -> {total, by_accommodation, name}
        for row in all_rows:
            if row["date"] != latest_date:
                continue
            code = row["lad_code"]
            if code not in la_totals:
                la_totals[code] = {"total": 0, "by_accommodation": {}, "name": row["la_name"]}
            la_totals[code]["total"] += row["people"]
            acc = row["accommodation"]
            la_totals[code]["by_accommodation"][acc] = la_totals[code]["by_accommodation"].get(acc, 0) + row["people"]

        # Also build trend (totals per date for Lancashire LAs)
        lancs_codes = set(v["ons"] for v in COUNCILS.values())
        trend_by_date = {}
        for row in all_rows:
            if row["lad_code"] not in lancs_codes:
                continue
            d = row["date"]
            if d not in trend_by_date:
                trend_by_date[d] = {}
            code = row["lad_code"]
            trend_by_date[d][code] = trend_by_date[d].get(code, 0) + row["people"]

        for code, data in la_totals.items():
            # Build per-LA trend from the last 4 dates
            la_trend = []
            for d in sorted(trend_by_date.keys())[-4:]:
                if code in trend_by_date[d]:
                    la_trend.append({"date": d, "people": trend_by_date[d][code]})

            asylum_data[code] = {
                "seekers_supported": data["total"],
                "by_accommodation": data["by_accommodation"],
                "local_authority_name": data["name"],
                "trend": la_trend,
                "latest_date": latest_date,
            }

        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
        print(f"  Parsed asylum data for {len(asylum_data)} local authorities", file=sys.stderr)

    except Exception as e:
        print(f"WARNING: Failed to fetch asylum data: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)

    return asylum_data


def fetch_resettlement_data(skip=False):
    """Download and parse Home Office resettlement data (Res_D01).

    Similar flat format: Date, Scheme, Region, Local Authority, LAD Code, People
    Returns dict keyed by ONS code: {total, by_scheme}
    """
    if skip:
        return {}

    try:
        import openpyxl
    except ImportError:
        return {}

    print("Downloading Home Office resettlement data...", file=sys.stderr)
    resettle_data = {}

    try:
        req = urllib.request.Request(RESETTLEMENT_URL, headers={"User-Agent": "AI-DOGE-ETL/1.0"})
        import tempfile
        with urllib.request.urlopen(req, timeout=120) as resp:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(resp.read())
                tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

        # Find the data sheet
        target_sheet = None
        for name in wb.sheetnames:
            lower = name.lower()
            if "data" in lower and ("res" in lower or "d01" in lower):
                target_sheet = name
                break
        if not target_sheet:
            for name in wb.sheetnames:
                lower = name.lower()
                if lower not in ("contents", "notes", "metadata", "cover", "cover_sheet", "list_of_fields"):
                    target_sheet = name
                    break

        if target_sheet:
            ws = wb[target_sheet]

            # First pass: find header row and columns
            header_row = None
            for i, row in enumerate(ws.iter_rows(values_only=True, max_row=5)):
                row_str = " ".join(str(c).lower() for c in row if c)
                if "local authority" in row_str or "lad code" in row_str:
                    header_row = i
                    headers = [str(c).strip().lower() if c else "" for c in row]
                    break

            if header_row is not None:
                # Find column indices
                code_col = next((j for j, h in enumerate(headers) if "lad code" in h or "code" in h), None)
                people_col = next((j for j, h in enumerate(headers) if "people" in h or "number" in h), None)
                scheme_col = next((j for j, h in enumerate(headers) if "scheme" in h), None)
                name_col = next((j for j, h in enumerate(headers) if "local authority" in h), None)

                if code_col is not None:
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i <= header_row:
                            continue
                        if not row or len(row) <= code_col:
                            continue
                        code = str(row[code_col]).strip() if row[code_col] else ""
                        if not code.startswith("E"):
                            continue

                        try:
                            people = int(float(row[people_col])) if people_col and len(row) > people_col and row[people_col] else 0
                        except (ValueError, TypeError):
                            people = 0

                        scheme = str(row[scheme_col]).strip() if scheme_col and len(row) > scheme_col and row[scheme_col] else "Unknown"

                        if code not in resettle_data:
                            resettle_data[code] = {"total": 0, "by_scheme": {}}
                        resettle_data[code]["total"] += people
                        resettle_data[code]["by_scheme"][scheme] = resettle_data[code]["by_scheme"].get(scheme, 0) + people

        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
        print(f"  Parsed resettlement data for {len(resettle_data)} local authorities", file=sys.stderr)

    except Exception as e:
        print(f"WARNING: Failed to fetch resettlement data: {e}", file=sys.stderr)

    return resettle_data


# ---- Main pipeline ----

def process_council(council_id, council_info, asylum_data, resettle_data):
    """Process a single council: fetch SNPP + merge asylum data."""
    ons_code = council_info["ons"]
    name = council_info["name"]

    print(f"Fetching SNPP for {name} ({ons_code})...", file=sys.stderr)

    try:
        projections = fetch_snpp(ons_code)
    except Exception as e:
        print(f"  ERROR: {e}", file=sys.stderr)
        return None

    if not projections:
        print(f"  WARNING: No projection data returned for {name}", file=sys.stderr)
        return None

    metrics = compute_metrics(projections)

    # Asylum data
    asylum = asylum_data.get(ons_code, {})
    resettle = resettle_data.get(ons_code, {})

    asylum_section = {"seekers_supported": 0}
    if asylum:
        asylum_section = {
            "seekers_supported": asylum.get("seekers_supported", 0),
            "by_accommodation": asylum.get("by_accommodation", {}),
            "trend": asylum.get("trend", []),
            "latest_date": asylum.get("latest_date", ""),
        }

    resettle_section = {"total": 0}
    if resettle:
        resettle_section = {
            "total": resettle.get("total", 0),
            "by_scheme": resettle.get("by_scheme", {}),
        }

    result = {
        "meta": {
            "source": "ONS 2022-based Sub-National Population Projections via Nomis API",
            "asylum_source": "Home Office Immigration Statistics, year ending December 2025",
            "council_id": council_id,
            "council_name": name,
            "ons_code": ons_code,
            "projection_base_year": "2022",
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
        },
        **metrics,
        "asylum": asylum_section,
        "resettlement": resettle_section,
    }

    return result


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Fetch ONS SNPP + Home Office asylum data")
    parser.add_argument("--council", help="Process single council")
    parser.add_argument("--skip-asylum", action="store_true", help="Skip asylum/resettlement data")
    parser.add_argument("--stdout", action="store_true", help="Print to stdout instead of files")
    args = parser.parse_args()

    councils_to_process = COUNCILS
    if args.council:
        if args.council not in COUNCILS:
            print(f"ERROR: Unknown council '{args.council}'", file=sys.stderr)
            print(f"Available: {', '.join(sorted(COUNCILS.keys()))}", file=sys.stderr)
            sys.exit(1)
        councils_to_process = {args.council: COUNCILS[args.council]}

    # Fetch asylum data once (shared across all councils)
    asylum_data = fetch_asylum_data(skip=args.skip_asylum)
    resettle_data = fetch_resettlement_data(skip=args.skip_asylum)

    summary = {"meta": {"generated": datetime.now().isoformat(), "source": "ONS SNPP 2022-based"}, "councils": {}}
    results = {}

    for council_id, info in councils_to_process.items():
        result = process_council(council_id, info, asylum_data, resettle_data)
        if result:
            results[council_id] = result

            # Add to summary
            summary["councils"][council_id] = {
                "name": info["name"],
                "population_2022": result["population_projections"].get("2022", 0),
                "population_2032": result["population_projections"].get("2032", 0),
                "population_2047": result["population_projections"].get("2047", 0),
                "growth_rate_pct": result["growth_rate_pct"],
                "dependency_2022": result["dependency_ratio_projection"].get("2022", 0),
                "dependency_2032": result["dependency_ratio_projection"].get("2032", 0),
                "working_age_2022": result["working_age_pct_projection"].get("2022", 0),
                "working_age_2032": result["working_age_pct_projection"].get("2032", 0),
                "asylum_seekers": result.get("asylum", {}).get("seekers_supported", 0),
                "resettlement_total": result.get("resettlement", {}).get("total", 0),
            }

        time.sleep(0.5)  # Polite rate limiting

    if args.stdout:
        json.dump(results, sys.stdout, indent=2)
        return

    # Write per-council files
    for council_id, result in results.items():
        out_path = DATA_DIR / council_id / "demographic_projections.json"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)
        print(f"  Written {out_path} ({out_path.stat().st_size:,} bytes)", file=sys.stderr)

    # Write shared summary
    summary_path = DATA_DIR / "shared" / "demographic_projections_summary.json"
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  Written {summary_path}", file=sys.stderr)

    print(f"\nDone! Processed {len(results)} councils.", file=sys.stderr)


if __name__ == "__main__":
    main()
